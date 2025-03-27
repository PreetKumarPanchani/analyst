import pandas as pd
import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from datetime import datetime

def get_metadata_and_data(file_path, sheet_name):
    """
    Extract metadata and data separately from an Excel file
    Returns metadata rows and a pandas dataframe with the actual data
    """
    try:
        # Read the entire sheet into a dataframe
        full_df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        # Find the row that contains "Sale ID" or similar header
        header_row = None
        sale_id_variations = ["Sale ID", "SaleID", "Sale_ID", "Sale Date"]
        
        for i, row in full_df.iterrows():
            # Convert row to string and check if any cell contains a header indicator
            row_as_str = row.astype(str).str.lower()
            if any(id_var.lower() in val for id_var in sale_id_variations for val in row_as_str):
                header_row = i
                break
        
        ## If header row is not found then take the first row where multiple values/columns are present , not just one column or value 
        if header_row is None:
            # Look for the first row with multiple non-empty values
            for i, row in full_df.iterrows():
                # Count non-empty cells in this row
                non_empty_cells = row.notna().sum()
                
                # If we have more than 2 non-empty cells, consider this a header row
                if non_empty_cells >= 3:
                    header_row = i
                    print(f"  Found potential header row at position {i+1} with {non_empty_cells} values in {file_path}, sheet {sheet_name}.")
                    break



        if header_row is None:
            # If we still couldn't find the header row, assume it's row 5 (common in these files)
            header_row = 5
            print(f"  Warning: Couldn't find header row in {file_path}, sheet {sheet_name}. Assuming row 6.")
        
        # Extract metadata (rows before the header)
        metadata_df = full_df.iloc[:header_row].copy()
        
        # Extract the actual data (including header row)
        data_df = full_df.iloc[header_row:].copy().reset_index(drop=True)
        
        # Set the first row as header
        data_df.columns = data_df.iloc[0]
        data_df = data_df.iloc[1:].reset_index(drop=True)
        
        # Add a column to track the source file
        data_df['Source_File'] = os.path.basename(file_path)
        
        # Extract company from filename
        company_match = re.search(r'(cpl|forge)_', os.path.basename(file_path).lower())
        if company_match:
            data_df['Company'] = company_match.group(1).upper()
        else:
            data_df['Company'] = "UNKNOWN"
            
        return metadata_df, data_df
    
    except Exception as e:
        print(f"Error processing {file_path}, sheet {sheet_name}: {e}")
        return pd.DataFrame(), pd.DataFrame()

def process_files(files, company, sheet_names, output_dir):
    """Process files for a specific company and generate merged outputs"""
    # Dictionaries to hold dataframes for each sheet
    all_metadata = {sheet: [] for sheet in sheet_names}
    all_data = {sheet: [] for sheet in sheet_names}
    
    # Process each file
    print(f"\nProcessing {len(files)} {company.upper()} files")
    for file in files:
        try:
            print(f"Reading file: {file}")
            for sheet in sheet_names:
                try:
                    metadata_df, data_df = get_metadata_and_data(file, sheet)
                    
                    if not data_df.empty:
                        # Convert all column names to string to avoid issues
                        data_df.columns = data_df.columns.astype(str)
                        
                        # Store metadata and data
                        all_metadata[sheet].append(metadata_df)
                        all_data[sheet].append(data_df)
                        print(f"  Extracted {len(data_df)} rows from {sheet}")
                    else:
                        print(f"  No data found in sheet {sheet}")
                        
                except Exception as e:
                    print(f"  Error processing sheet {sheet} in {file}: {e}")
        except Exception as e:
            print(f"Error reading file {file}: {e}")
    
    # Create merged files for each sheet
    for sheet in sheet_names:
        if not all_data[sheet]:
            print(f"No data found for {sheet} in any {company} file")
            continue
            
        try:
            # Get example metadata from the first file
            example_metadata = all_metadata[sheet][0] if all_metadata[sheet] else None
            
            # Merge all data for this sheet
            # Find all unique columns across all dataframes
            all_columns = set()
            for df in all_data[sheet]:
                all_columns.update(df.columns)
            
            # Ensure all dataframes have all columns
            for i, df in enumerate(all_data[sheet]):
                missing_cols = all_columns - set(df.columns)
                if missing_cols:
                    for col in missing_cols:
                        df[col] = np.nan
            
            # Concatenate all data
            merged_data = pd.concat(all_data[sheet], ignore_index=True)
            
            # Create output Excel file
            output_file = os.path.join(output_dir, f"{company}_{sheet.replace(' ', '_')}.xlsx")
            
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            
            # Add metadata if available
            if example_metadata is not None:
                # Update the date range in metadata
                # Find "From:" and "To:" rows
                from_row = None
                to_row = None
                for i, row in example_metadata.iterrows():
                    row_str = row.astype(str).str.lower()
                    if any('from:' in val for val in row_str):
                        from_row = i
                    if any('to:' in val for val in row_str):
                        to_row = i
                
                # Update metadata with combined date range
                try:
                    if from_row is not None and to_row is not None:
                        # Find earliest "From" date across all files
                        earliest_from = None
                        latest_to = None
                        
                        for meta_df in all_metadata[sheet]:
                            try:
                                from_text = meta_df.iloc[from_row, 1] if from_row < len(meta_df) else None
                                to_text = meta_df.iloc[to_row, 1] if to_row < len(meta_df) else None
                                
                                if from_text and isinstance(from_text, str) and 'AM' in from_text or 'PM' in from_text:
                                    try:
                                        # Extract date from string like "01/01/2024 01:00 AM"
                                        from_date = datetime.strptime(from_text, "%d/%m/%Y %I:%M %p")
                                        if earliest_from is None or from_date < earliest_from:
                                            earliest_from = from_date
                                    except:
                                        pass
                                
                                if to_text and isinstance(to_text, str) and 'AM' in to_text or 'PM' in to_text:
                                    try:
                                        # Extract date from string like "29/02/2024 11:00 PM"
                                        to_date = datetime.strptime(to_text, "%d/%m/%Y %I:%M %p")
                                        if latest_to is None or to_date > latest_to:
                                            latest_to = to_date
                                    except:
                                        pass
                            except:
                                continue
                        
                        # Update the metadata with the new date range
                        if earliest_from is not None:
                            example_metadata.iloc[from_row, 1] = earliest_from.strftime("%d/%m/%Y %I:%M %p")
                        if latest_to is not None:
                            example_metadata.iloc[to_row, 1] = latest_to.strftime("%d/%m/%Y %I:%M %p")
                except Exception as e:
                    print(f"  Warning: Couldn't update date range in metadata: {e}")
                
                # Add metadata to worksheet
                for r_idx, row in example_metadata.iterrows():
                    for c_idx, value in enumerate(row):
                        if pd.notna(value):  # Only add non-NA values
                            ws.cell(row=r_idx+1, column=c_idx+1, value=value)
                
                # Add an empty row after metadata
                header_row_index = len(example_metadata) + 1
            else:
                header_row_index = 1
            
            # Add headers and data
            # First add the headers
            for c_idx, col_name in enumerate(merged_data.columns):
                if col_name not in ['Source_File', 'Company']:  # Skip internal tracking columns
                    ws.cell(row=header_row_index, column=c_idx+1, value=col_name)
            
            # Then add the data rows
            for r_idx, row in merged_data.iterrows():
                for c_idx, col_name in enumerate(merged_data.columns):
                    if col_name not in ['Source_File', 'Company']:  # Skip internal tracking columns
                        ws.cell(row=header_row_index+r_idx+1, column=c_idx+1, value=row[col_name])
            
            # Save the workbook
            wb.save(output_file)
            print(f"Created {output_file} with {len(merged_data)} rows")
            
        except Exception as e:
            print(f"Error creating merged file for {sheet}: {e}")
    
    return all_data

def verify_merge(all_data, company):
    """Verify that data from all files was merged correctly"""
    print(f"\nVerification for {company.upper()}:")
    
    for sheet, data_list in all_data.items():
        if not data_list:
            print(f"No data found for {sheet}")
            continue
        
        total_rows = sum(len(df) for df in data_list)
        unique_files = set()
        
        for df in data_list:
            unique_files.update(df['Source_File'].unique())
        
        print(f"{sheet}: Merged data from {len(unique_files)} files, total rows: {total_rows}")
        print(f"  Source files: {', '.join(sorted(unique_files))}")


# Main function
def main():
    dir_path = r'C:\Users\preet\Downloads\Liquid_SCM\Code\sales_forecast\backend\data\raw'

    # Define files
    cpl_files = ["cpl_1-2_export_transactions_report_20250222_195204.xlsx",
                 "cpl_3-5_export_transactions_report_20250222_195244.xlsx",
                 "cpl_6-8_export_transactions_report_20250222_195343.xlsx", 
                 "cpl_9-11_export_transactions_report_20250222_195434.xlsx",
                 "cpl_12-14_export_transactions_report_20250222_195503.xlsx",
                 ]

    forge_files = [
                   "forge_3-4_export_transactions_report_20250222_194908.xlsx",
                   "forge_1-2_export_transactions_report_20250222_195117.xlsx",
                   "forge_5-6_export_transactions_report_20250222_194829.xlsx",
                   "forge_7-8_export_transactions_report_20250222_194803.xlsx",
                   "forge_9-10_export_transactions_report_20250222_194731.xlsx",
                   "forge_11-12_export_transactions_report_20250222_194658.xlsx",
                   "forge_13-14_export_transactions_report_20250222_194607.xlsx"]

    # Define sheet names
    sheet_names = ['Sales', 'Sales Items', 'Sales Payments', 'Sales Refunds', 'Deleted Sales Items']
    
    # Add the output directory to the file paths
    cpl_files = [os.path.join(dir_path, f) for f in cpl_files]
    forge_files = [os.path.join(dir_path, f) for f in forge_files]

    # Create output directory
    output_dir = "merged_dataset"
    os.makedirs(output_dir, exist_ok=True)
    

    # Process CPL files
    cpl_data = process_files(cpl_files, "cpl", sheet_names, output_dir)
    
    # Process Forge files
    forge_data = process_files(forge_files, "forge", sheet_names, output_dir)
    
    # Verify the merges
    verify_merge(cpl_data, "cpl")
    verify_merge(forge_data, "forge")
    
    print("\nMerging completed successfully.")
    print(f"The merged files can be found in the '{output_dir}' directory.")



if __name__ == "__main__":
    main()